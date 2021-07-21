import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
driver = webdriver.Chrome()
driver.get('https://academic.oup.com/journals/search-results?page=1&q=data+mining&SearchSourceType=1&allJournals=1')
xlpath = "./task.xlsx"
workbook = openpyxl.load_workbook(xlpath)
sheet = workbook.active
title_LINKS = '//*[@id="ContentColumn"]/div[number]/h4/a'
sheet.cell(row=1,column=1).value = "TITLE"
sheet.cell(row=1,column=2).value = "AUTHOR"
sheet.cell(row=1,column=3).value = "YEAR"
sheet.cell(row=1,column=4).value = "LINK"

for i in range(11,31):
    p = title_LINKS.replace('number',str(i))
    data = driver.find_element_by_xpath(p)
    sheet.cell(row=i-9,column=1).value = data.text
    sheet.cell(row=i-9,column=4).value = data.get_attribute("href")
count = 1 
for i in driver.find_elements_by_xpath('//div[@class="sr-list al-article-box al-normal clearfix"]'):
    author = i.find_elements_by_xpath('.//span[@class="wi-fullname brand-fg"]/a')
    authors = []
    for xt in author:
        authors.append(xt.text)
    sheet.cell(row=count+1,column=2).value = ','.join(authors)
    count = count + 1

for i in range(1,19):
    year = driver.find_elements_by_class_name('al-pub-date')
    sheet.cell(row=i+1,column=3).value = year[i].text

for r in range(1,11):
    print(r," :data")
    for c in range(1,5):
        if c == 1: print("TITLE:")
        if c == 2: print("AUTHORS:")
        if c == 3: print("YEAR:")
        if c == 4: print("LINK:")
        print(sheet.cell(row=r,column=c).value)
workbook.save(xlpath)
